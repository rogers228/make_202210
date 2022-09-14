if True: # 固定引用開發環境 或 發佈環境 的 路徑
    import os, sys, custom_path
    config_path = os.getcwd() if os.getenv('COMPUTERNAME')=='VM-TESTER' else custom_path.custom_path['make_202210'] # 目前路徑
    sys.path.append(config_path)

import os
import openpyxl
from openpyxl.comments import Comment #註解
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D #插入圖片用
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker #插入圖片用
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU #插入圖片用
from openpyxl.drawing.image import Image #插入圖片用
from PIL import Image as pil_image
from openpyxl.utils import get_column_letter #轉換
from tool_style import *
from config import *

class tool_excel(): #讀取excel 單一零件
    def __init__(self, file, workbook, sh):
        self.file = file
        self.workbook = workbook
        self.sh = sh # excel sheet

    def c_write(self, row, column, value = '', font = font_9, alignment = ah_left, border = no_border, fillcolor = cf_none):
        #寫入儲存格 並設定格式
        cell = self.sh.cell(row, column)
        cell.value = value
        if font:
            cell.font = font
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border
        if fillcolor:
            cell.fill = fillcolor

    def c_merge(self, start_row, start_column, end_row, end_column):
        self.sh.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column) #合併儲存格

    def set_page_layout(self): # 頁面設定layout
        cm2in = lambda x: x/2.54
        self.sh.page_margins = openpyxl.worksheet.page.PageMargins(
            left=cm2in(1.2),
            right=cm2in(1.2),
            top=cm2in(1.0),
            bottom=cm2in(0.5),
            header=cm2in(1.0),
            footer=cm2in(0.5))

    def set_page_layout_horizontal(self): # 頁面設定layout  橫式
        cm2in = lambda x: x/2.54
        self.sh.page_margins = openpyxl.worksheet.page.PageMargins(
            left=cm2in(0.5),
            right=cm2in(0.5),
            top=cm2in(0.5),
            bottom=cm2in(0.5),
            header=cm2in(1.0),
            footer=cm2in(0.5))

        self.sh.page_setup.paperSize = self.sh.PAPERSIZE_A4
        # Paper size 紙張大小
            # PAPERSIZE_LETTER = '1'
            # PAPERSIZE_LETTER_SMALL = '2'
            # PAPERSIZE_TABLOID = '3'
            # PAPERSIZE_LEDGER = '4'
            # PAPERSIZE_LEGAL = '5'
            # PAPERSIZE_STATEMENT = '6'
            # PAPERSIZE_EXECUTIVE = '7'
            # PAPERSIZE_A3 = '8'
            # PAPERSIZE_A4 = '9'
            # PAPERSIZE_A4_SMALL = '10'
            # PAPERSIZE_A5 = '11'
        self.sh.page_setup.orientation = self.sh.ORIENTATION_LANDSCAPE
        # Page orientation 紙張方向
            # ORIENTATION_PORTRAIT = 'portrait' #縱向
            # ORIENTATION_LANDSCAPE = 'landscape' #横向
        self.sh.sheet_view.zoomScale = 100 # 檢視縮放
        self.sh.page_setup.scale = 75      # 列印縮放比例
        self.sh.print_options.horizontalCentered=True # 水平居中

    def c_image(self, row, column, imgPath, width, height, rowoffset=0, coloffset=0): #插入圖片
        # imgPath 圖片路徑  請在程序外先檢查是否存在
        img = Image(imgPath)
        img.width = width
        img.height = height
        cell_h_to_EMU = lambda h: cm_to_EMU((h * 49.77)/99)         # cell height EMU單位
        cell_w_to_EMU = lambda w: cm_to_EMU((w * (18.65-1.71))/10)  # cell width  EMU單位
        coloffset = cell_w_to_EMU(coloffset) #偏移
        rowoffset = cell_h_to_EMU(rowoffset) #偏移
        marker = AnchorMarker(col=column-1, colOff=coloffset, row=row-1, rowOff=rowoffset) #建立標記位置  由1始
        size = XDRPositiveSize2D(pixels_to_EMU(width), pixels_to_EMU(height))
        img.anchor = OneCellAnchor(_from=marker, ext=size) #img 定位
        self.sh.add_image(img)

    def c_image2(self, row, column, pdno, rowoffset=0, coloffset=0, max_height=80): #插入圖片
        imgfullpath = (os.path.join(config_image_dir, f'{pdno}.bmp'))
        # print(imgfullpath)
        if not os.path.isfile(imgfullpath): # 無此檔案
            return

        # step 1 求縮圖尺寸 
        max_width  = 80
        f_height = lambda w,h: int((max_width*h)/w) # 依照 max_width 求等比 height
        f_width = lambda w,h: int((max_height*w)/h) # 依照 max_height 求等比 width
        (w, h) = pil_image.open(imgfullpath).size
        w, h = max_width, f_height(w,h) # 依照 max_width 求等比 height
        if h > max_height:
            w, h = f_width(w,h), max_height # 依照 max_width 求等比 height
        width, height = w, h

        # step 2 插入圖片
        img = Image(imgfullpath)
        img.width = width
        img.height = height
        cell_h_to_EMU = lambda h: cm_to_EMU((h * 49.77)/99)         # cell height EMU單位
        cell_w_to_EMU = lambda w: cm_to_EMU((w * (18.65-1.71))/10)  # cell width  EMU單位
        coloffset = cell_w_to_EMU(coloffset) #偏移
        rowoffset = cell_h_to_EMU(rowoffset) #偏移
        marker = AnchorMarker(col=column-1, colOff=coloffset, row=row-1, rowOff=rowoffset) #建立標記位置  由1始
        size = XDRPositiveSize2D(pixels_to_EMU(width), pixels_to_EMU(height))
        img.anchor = OneCellAnchor(_from=marker, ext=size) #img 定位
        self.sh.add_image(img)

    def c_column_width(self, width_list): # 設定欄寬
        for i in range(len(width_list)):
            self.sh.column_dimensions[get_column_letter(i+1)].width = width_list[i]

    def c_row_height(self, row_index, row_height): # 設定列高
        self.sh.row_dimensions[row_index].height = row_height 

    def c_line_border(self, row, start_column, columns, border = bottom_border): #畫格線
        for i in range(start_column, start_column + columns):
            self.sh.cell(row, i).border = border #style

    def c_comm(self, row, column, message):
        #註解
        comment = Comment(message, "Author")
        self.sh.cell(row, column).comment = comment

    def c_fill(self, row, column, fillcolor = cf_yellow):
        #填充顏色
        self.sh.cell(row, column).fill = fillcolor

    def save_xls(self): # 儲存
        try:
            self.workbook.save(self.file) #save
        except:
            print('儲存時發生錯誤，無法處理該檔案，有可能檔案已被開啟尚未關閉!')

    def open_xls(self):
        if os.path.exists(self.file): #檔案存在
            # 使用cmd 使用excel啟動 最大化 該檔案
            cmd = r'start "" /max EXCEL.EXE "' + self.file + '"'
            # print(cmd)
            os.system(cmd)

def test1():
    print('test1')

if __name__ == '__main__':
    test1()
    print('ok')