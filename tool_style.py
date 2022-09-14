if True: # 固定引用開發環境 或 發佈環境 的 路徑
    import os, sys, custom_path
    config_path = os.getcwd() if os.getenv('COMPUTERNAME')=='VM-TESTER' else custom_path.custom_path['make_202210'] # 目前路徑
    sys.path.append(config_path)
    
from openpyxl.styles import Alignment, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill

if True: # border
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    bottom_border = Border(bottom=Side(style='thin'))
    top_border = Border(top=Side(style='thin'))
    bt_border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
    bottom_border_sk = Border(bottom=Side(style='dotted'))
    none_border = Side(border_style=None)
    no_border = Border(
        left = none_border, 
        right = none_border, 
        top = none_border,
        bottom = none_border,
    )

if True: # font
    # font = Font(name='Calibri', size=11, bold=False,
    # ...                 italic=False,
    # ...                 vertAlign=None,
    # ...                 underline='none',
    # ...                 strike=False,
    # ...                 color='FF000000')
    font_8 = Font(size = "8")
    font_A_8 = Font(name = "Arial" ,size = "8")
    font_9 = Font(size = "9")
    font_9B = Font(size = "9", bold=True)
    font_9B_Gadugi = Font(size = "9", bold=True, name='Gadugi')
    font_10 = Font(size = "10")
    font_A_10 = Font(name = "Arial" ,size = "10")
    font_A_10I = Font(name = "Arial" ,size = "10", italic=True)
    font_A_10G = Font(name = "Arial" ,size = "10", color='cccccc')
    font_10B = Font(size = "10", bold=True)
    font_10_Calibri = Font(size = "10", name='Calibri')
    font_10_Calibri_g = Font(size = "10", name='Calibri', color='cccccc')
    font_11 = Font(size = "11")
    font_11_Calibri = Font(size = "11", name='Calibri')
    font_11_Calibri_green = Font(size = "11", name='Calibri', color='339933')
    font_11_green = Font(size = "11", color='339933')
    font_11B = Font(size = "11", bold=True)
    font_11B_link = Font(size = "11", bold=True, color='0212BC')
    font_11B_link_ab = Font(size = "11", bold=True, color='C70BB3')
    font_11B_link_pdm = Font(size = "11", bold=True, color='00751E')
    font_12 = Font(size = "12")
    font_14BI = Font(size = "14", bold=True, italic=True)
    font_16I = Font(size = "16", italic=True)
    font_18B = Font(size = "18", bold=True)
    font_26BI = Font(size = "26", bold=True, italic=True)
    font_9_page = Font(name='Calibri', size=9, italic=True)
    font_10_pub_time = Font(size = "10", name='Calibri')
    font_12_pub_page = Font(size = "12", name='Calibri')

if True: # Alignment
    #水平方向的對齊方式
    #horizontal value must be one of {‘justify’, ‘centerContinuous’, ‘right’, ‘center’, ‘general’, ‘fill’, ‘distributed’, ‘left’}
    ah_left = Alignment(horizontal='left')
    ah_left_top = Alignment(horizontal='left', vertical ='top')
    ah_left_bottom = Alignment(horizontal='left', vertical ='bottom')
    ah_left_center = Alignment(horizontal='left', vertical ='center')
    ah_center = Alignment(horizontal='center')
    ah_center_bottom = Alignment(horizontal='center', vertical ='bottom')
    ah_right = Alignment(horizontal='right')
    ah_right_bottom = Alignment(horizontal='right', vertical ='bottom')
    ahv_center = Alignment(horizontal='center',vertical='center')

    ah_wr = Alignment(wrapText=True)  #自動換行
    ah_top_wr = Alignment(vertical='top', wrapText=True)  #自動換行
    ah_center_top = Alignment(horizontal='center', vertical ='top')
    ah_center_bottom = Alignment(horizontal='center', vertical ='bottom')
    
if True: # color fill 填滿顏色
    cf_none = PatternFill(fill_type=None) #不填滿
    cf_yellow = PatternFill(start_color="FFFF00",  fill_type = "solid")  # 黃色
    cf_gray = PatternFill(start_color="CACACA",  fill_type = "solid")    # 灰色
    cf_blue = PatternFill(start_color="3CCEFF",  fill_type = "solid")    # 藍色
    cf_khaki = PatternFill(start_color="F0B000",  fill_type = "solid")    # 土黃色
    cf_green = PatternFill(start_color="C3F069",  fill_type = "solid")    # 綠色    