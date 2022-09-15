if True: # 固定引用開發環境 或 發佈環境 的 路徑
    import os, sys, custom_path
    config_path = os.getcwd() if os.getenv('COMPUTERNAME')=='VM-TESTER' else custom_path.custom_path['make_202210'] # 目前路徑
    sys.path.append(config_path)

import time
import openpyxl
import pandas as pd
import PySimpleGUI as sg
from tool_excel2 import tool_excel
from tool_style import *
import tool_file
import tool_db_make
from config import *

class Report_bng01(tool_excel):
    def __init__(self, filename, department):
        self.fileName = filename
        self.department = department # 部門id
        self.report_name = 'bng01' # 製造排程表
        self.report_dir = config_report_dir # 資料夾名稱
        self.report_path = os.path.join(os.path.expanduser(r'~\Documents'), self.report_dir) #資料夾路徑
        self.file_tool = tool_file.File_tool() # 檔案工具並初始化資料夾
        if self.report_path is None:
            print('找不到路徑')
            sys.exit() #正式結束程式
        self.file_tool.clear(self.report_name) # 清除舊檔
        self.db =  tool_db_make.db_make()
        self.comp_base(); self.comp_base2()
        self.comp_data()  # self.df_bn
        self.comp_data2() # self.df_fbnr
        self.create_excel()
        self.output()
        self.save_xls()
        self.open_xls()

    def create_excel(self):
        self.wb = openpyxl.Workbook()
        wb = self.wb
        self.df_ma = self.db.get_ma_df(self.department) # 所有機台
        for i, r in self.df_ma.iterrows():
            wb.create_sheet(r['ma008'])

        self.dic_ws = {1: '未排程',5: '待排程',6: '建議外包'} # 未排程資料
        for k, v in self.dic_ws.items():
            wb.create_sheet(v)

        wb.remove(wb['Sheet'])
        self.xlsfile = os.path.join(self.report_path, self.fileName)
        wb.save(filename = self.xlsfile)

    def comp_base(self):
        # 基礎設定
        # name, width, sql_column_name
        lis_base = []; a = lis_base.append
        a('順序碼,     7, bn005') 
        a('工程名稱,   25, pm002')
        a('派工數量,    7, br008') 
        a('預計生產數量,7, bn006') 
        a('材料,      10, fbn023') 
        a('加工時間,   23, bn010') 
        a('預計開始,   10, bn007') 
        a('預計完工,   10, bn008') 
        a('實際開始,   10, bn044') 
        a('實際完工,   10, bn045') 
        a('製令編號,   17, sbr003') 
        a('材料規格,   25, sbt004') 
        a('預交日,     10, br009') 
        a('急緩等級,   10, fbr019') 
        a('品名,       20, br006') 
        a('品號,       12, br005') 
        a('規格,       17, br007') 
        a('良品數,     7,  bn042') 
        a('狀態,       8, fbn061') 
        a('排程ID,     7,  bn001') 
        a('派工ID,     7,  bn003') 
        a('庫存數量,    7, br025') 
        a('庫存查詢時間,16, br026') 
        a('校機人員,    9,  bn070') 
        a('操作人員,    9,  bn071') 

        lis_e1, lis_e2, lis_e3 = [],[],[]
        for e in lis_base:
            [e1, e2, e3]= e.split(',')
            lis_e1.append(e1.strip())
            lis_e2.append(int(e2.strip()))
            lis_e3.append(e3.strip())
        self.xls_index =dict(zip(lis_e1, [e for e in range(1,len(lis_e1)+1)]))
        self.xls_width =dict(zip(lis_e1, lis_e2))
        self.xls_sqlcn =dict(zip(lis_e1, lis_e3))

    def comp_base2(self):
        # 基礎設定 第二部分
        # name, width, sql_column_name
        lis_base = []; a = lis_base.append
        a('派工ID,   7, br001') 
        a('製令單別, 7, br002')
        a('製令單號,12, br003') 
        a('品號,    12, br005') 
        a('品名,    30, br006') 
        a('規格,    30, br007') 
        a('製程,    16, br004') 
        a('派工數量,  7, br008') 
        a('預交日,   12, br009') 
        a('急緩等級, 10, fbr019') 
        lis_e1, lis_e2, lis_e3 = [],[],[]
        for e in lis_base:
            [e1, e2, e3]= e.split(',')
            lis_e1.append(e1.strip())
            lis_e2.append(int(e2.strip()))
            lis_e3.append(e3.strip())
        self.xls_index_2 =dict(zip(lis_e1, [e for e in range(1,len(lis_e1)+1)]))
        self.xls_width_2 =dict(zip(lis_e1, lis_e2))
        self.xls_sqlcn_2 =dict(zip(lis_e1, lis_e3))

    def comp_data(self): # 資料加工
        sy002 = self.db.get_sy002() # 顯示時間
        df = self.db.get_bn_df(self.department, sy002) # 所有排程資料 
        date_columns = ['bn007','bn008','bn044','bn045','br009','br026']
        df[date_columns] = df[date_columns].fillna(value='') # 填充空白
        df[date_columns] = df[date_columns].astype(str) # 時間轉文字
        int_columns = ['bn023','br019','bn061','bn062','bn066']
        df[int_columns] = df[int_columns].fillna(value=0) # 填充
        df.insert(len(df.columns), 'fbn023', ['']*len(df.index), True) #插入欄
        df.insert(len(df.columns), 'fbr019', ['']*len(df.index), True) #插入欄
        df.insert(len(df.columns), 'fbn061', ['']*len(df.index), True) #插入欄
        df.loc[df.index[df.bn023==1][:],'fbn023']='●來料'
        df.loc[df.index[df.bn023==0][:],'fbn023']='○缺料'
        df.loc[df.index[df.br019==1][:],'fbr019']='1不急'
        df.loc[df.index[df.br019==2][:],'fbr019']='2普通'
        df.loc[df.index[df.br019==3][:],'fbr019']='3急'
        df.loc[df.index[df.br019==4][:],'fbr019']='4特級'
        df.loc[df.index[df.br019==5][:],'fbr019']='5插單'
        df.loc[df.index[df.bn061==0][:],'fbn061']='0未開始'
        df.loc[df.index[df.bn061==1][:],'fbn061']='1準備中'
        df.loc[df.index[df.bn061==2][:],'fbn061']='2準備好'
        df.loc[df.index[df.bn061==3][:],'fbn061']='3校模中'
        df.loc[df.index[df.bn061==4][:],'fbn061']='4生產中'
        df.loc[df.index[df.bn061==5][:],'fbn061']='5已完工'
        self.df_bn = df

    def comp_data2(self): # 資料加工 第二部分
        df = self.db.get_fbnr_df(self.department) # #所有尚未排程 的派工資料)
        date_columns = ['br009']
        df[date_columns] = df[date_columns].fillna(value='') # 填充空白
        df[date_columns] = df[date_columns].astype(str) # 時間轉文字
        int_columns = ['br019']
        df[int_columns] = df[int_columns].fillna(value=0) # 填充
        df.insert(len(df.columns), 'fbr019', ['']*len(df.index), True) #插入欄
        df.loc[df.index[df.br019==1][:],'fbr019']='1不急'
        df.loc[df.index[df.br019==2][:],'fbr019']='2普通'
        df.loc[df.index[df.br019==3][:],'fbr019']='3急'
        df.loc[df.index[df.br019==4][:],'fbr019']='4特級'
        df.loc[df.index[df.br019==5][:],'fbr019']='5插單'
        self.df_fbnr = df

    def output(self):
        if True: # style, func
            f10g =font_10_Calibri_g
            f10 = font_10_Calibri
            f11 = font_11_Calibri
            f11gr = font_11_Calibri_green
            # fill color
            fill_color_bn061 = {
                0: cf_none, # 0未開始
                1: cf_none, # 1準備中
                2: cf_none, # 2準備好
                3: cf_none, # 3校模中
                4: cf_green, # 4生產中
                5: cf_blue} # 5已完工

            fill_color_bn066 = {
                1: cf_purple} #  1插單

            fill_color_bn062 = {
                1: cf_red} #  1.異常停機

            # func, method
            write=self.c_write; fill=self.c_fill; comm=self.c_comm; img=self.c_image2; column_w=self.c_column_width

        wb = self.wb
        # 第1部分 機台排程狀況
        df_bn = self.df_bn
        df_ma = self.df_ma
        x_index = self.xls_index
        x_width = self.xls_width
        x_sqlcn = self.xls_sqlcn
        for mai, mar in self.df_ma.iterrows():
            sh = wb[mar['ma008']]
            super().__init__(self.xlsfile, wb, sh) # 傳遞引數給父class
            cr=1; column_w(list(x_width.values())) # 設定欄寬
            for name, index in x_index.items():
                write(cr, index, name, f11, alignment=ah_wr, border=bt_border, fillcolor=cf_gray) # 欄位名稱

            df_w = df_bn.loc[df_bn['bn002']==mar['ma001']]
            for i, r in df_w.iterrows():
                cr+=1

                fc = fill_color_bn061[r['bn061']] # fill color
                for name, index in x_index.items():
                    scn = x_sqlcn[name] # sql_column_name
                    write(cr, index, r[scn], f11, border=bottom_border, fillcolor=fc)

                if r['bn066']==1: #插單
                    fc = fill_color_bn066[r['bn066']] # fill color
                    for cj in range(2, len(x_index)+1):
                        fill(cr, cj, fillcolor=fc)

                if r['bn062']==1: #異常停機
                    fc = fill_color_bn066[r['bn062']] # fill color
                    for cj in range(2, len(x_index)+1):
                        fill(cr, cj, fillcolor=fc)

        # 第2部分 尚未排程的 派工資料
        x_index = self.xls_index_2
        x_width = self.xls_width_2
        x_sqlcn = self.xls_sqlcn_2
        df_fbnr = self.df_fbnr
        for wsid, sheet_name in self.dic_ws.items():
            sh = wb[sheet_name]
            super().__init__(self.xlsfile, wb, sh) # 傳遞引數給父class
            cr=1; column_w(list(x_width.values())) # 設定欄寬
            for name, index in x_index.items():
                write(cr, index, name, f11, alignment=ah_wr, border=bt_border, fillcolor=cf_gray) # 欄位名稱

            df_w = df_fbnr.loc[df_fbnr['br015']==wsid]
            for i, r in df_w.iterrows():
                cr+=1
                for name, index in x_index.items():
                    scn = x_sqlcn[name] # sql_column_name
                    write(cr, index, r[scn], f11, border=bottom_border)

def test1():
    fileName = 'bng01' + '_' + time.strftime("%Y%m%d%H%M%S", time.localtime()) + '.xlsx'
    Report_bng01(fileName, 1)
    print('ok')

if __name__ == '__main__':
    test1()
    sys.exit() #正式結束程式